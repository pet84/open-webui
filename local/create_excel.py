import openpyxl
from openpyxl import Workbook

# Vytvoření nového sešitu
wb = Workbook()
ws = wb.active
ws.title = "Přehled kategorií"

# Hlavičky
ws['A1'] = 'Kategorie'
ws['B1'] = '2024'
ws['C1'] = '2025'

# Data
data = [
    ('PENB BD + ostatní', 325, 394),
    ('PENB RD s projektem', 158, 78),
    ('PENB RD se zaměřením', 87, 65),
    ('Projekt RD', 64, 62),
    ('Pasport', 15, 10),
]

# Přidání dat do tabulky
for row_idx, (kategorie, hodnota_2024, hodnota_2025) in enumerate(data, start=2):
    ws[f'A{row_idx}'] = kategorie
    ws[f'B{row_idx}'] = hodnota_2024
    ws[f'C{row_idx}'] = hodnota_2025

# Formátování hlaviček
from openpyxl.styles import Font, PatternFill, Alignment

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Nastavení šířky sloupců
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 10

# Uložení souboru
filename = 'prehled_kategorii_2024_2025.xlsx'
wb.save(filename)
print(f"Excel soubor '{filename}' byl úspěšně vytvořen!")
